"""
generador_excel.py – Genera la liquidación de sueldo en formato Excel (.xlsx).
Produce una hoja profesional con formato y fórmulas.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from datetime import datetime


# ─── Colores hex (ARGB) ──────────────────────
C_VERDE_OSC   = "FF1B4332"
C_VERDE_MED   = "FF2D6A4F"
C_VERDE_CLA   = "FF52B788"
C_GRIS_CLARO  = "FFF1F5F2"
C_ROJO        = "FFDC3545"
C_AMARILLO    = "FFFFC107"
C_BLANCO      = "FFFFFFFF"
C_NEGRO       = "FF212529"
C_GRIS_TEXTO  = "FF6C757D"


def _borde(tipo="thin"):
    lado = Side(style=tipo)
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _fmt_clp(monto: int) -> str:
    return f"${monto:,.0f}".replace(",", ".")


def _celda(ws, fila, col, valor, negrita=False, alinear="left",
           bg=None, color_texto=None, formato_num=None, tamanio=10):
    c = ws.cell(row=fila, column=col, value=valor)
    c.font = Font(name="Calibri", bold=negrita, size=tamanio,
                  color=color_texto or C_NEGRO)
    c.alignment = Alignment(horizontal=alinear, vertical="center", wrap_text=True)
    c.border = _borde()
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if formato_num:
        c.number_format = formato_num
    return c


def generar_excel_liquidacion(datos: dict, resumen: dict, ruta_salida: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Liquidación"

    # ─── Ancho de columnas ────────────────────
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16

    fila = 1

    # ─── ENCABEZADO ───────────────────────────
    ws.row_dimensions[fila].height = 35
    ws.merge_cells(f"A{fila}:D{fila}")
    c = ws.cell(row=fila, column=1,
                value="🌿  SERVICARG – LIQUIDACIÓN DE SUELDO")
    c.font = Font(name="Calibri", bold=True, size=16, color=C_BLANCO)
    c.fill = PatternFill("solid", fgColor=C_VERDE_OSC)
    c.alignment = Alignment(horizontal="center", vertical="center")
    fila += 1

    ws.row_dimensions[fila].height = 20
    ws.merge_cells(f"A{fila}:D{fila}")
    c = ws.cell(row=fila, column=1,
                value=f"{datos['mes']} {datos['anio']}  |  Emitido: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c.font = Font(name="Calibri", size=10, color=C_VERDE_CLA)
    c.fill = PatternFill("solid", fgColor=C_VERDE_OSC)
    c.alignment = Alignment(horizontal="center", vertical="center")
    fila += 1

    fila += 1  # espacio

    # ─── DATOS DEL TRABAJADOR ─────────────────
    def encabezado_seccion(titulo, color_bg=C_VERDE_MED):
        nonlocal fila
        ws.row_dimensions[fila].height = 20
        ws.merge_cells(f"A{fila}:D{fila}")
        c = ws.cell(row=fila, column=1, value=titulo)
        c.font = Font(name="Calibri", bold=True, size=11, color=C_BLANCO)
        c.fill = PatternFill("solid", fgColor=color_bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _borde()
        fila += 1

    def fila_info(label, valor):
        nonlocal fila
        ws.row_dimensions[fila].height = 18
        _celda(ws, fila, 1, label, negrita=True, bg=C_GRIS_CLARO, tamanio=9)
        ws.merge_cells(f"B{fila}:D{fila}")
        c = ws.cell(row=fila, column=2, value=valor)
        c.font = Font(name="Calibri", size=9)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _borde()
        fila += 1

    encabezado_seccion("DATOS DEL TRABAJADOR")
    fila_info("Nombre completo", datos["nombre"])
    fila_info("RUT", datos["rut"] or "—")
    fila_info("Empresa / Faena", datos["empresa"] or "—")
    fila_info("Labor / Cargo", datos["labor"])
    fila_info("Sueldo base", _fmt_clp(datos["sueldo_base"]))
    fila_info("Días trabajados", str(datos["dias_trabajados"]))
    fila_info("Horas extras", str(datos.get("horas_extra", 0)))

    fila += 1  # espacio

    # ─── HABERES ──────────────────────────────
    encabezado_seccion("HABERES")

    # Sub-encabezado columnas
    ws.row_dimensions[fila].height = 18
    for col, txt in enumerate(["Concepto", "Imponible", "No Imponible", "Total"], 1):
        _celda(ws, fila, col, txt, negrita=True,
               alinear="center", bg="FFD4EDDA", tamanio=9)
    fila += 1

    def fila_haber(concepto, imponible, no_imponible, total, bold=False):
        nonlocal fila
        ws.row_dimensions[fila].height = 17
        bg = C_BLANCO if fila % 2 == 0 else C_GRIS_CLARO
        _celda(ws, fila, 1, concepto, negrita=bold, bg=bg, tamanio=9)
        for col, val in enumerate([imponible, no_imponible, total], 2):
            c_cell = _celda(ws, fila, col, val,
                            negrita=bold, alinear="right",
                            bg=bg, tamanio=9,
                            formato_num='#,##0')
        fila += 1

    fila_haber("Sueldo Base (proporcional a días trabajados)",
               resumen["sueldo_base_proporcional"], 0,
               resumen["sueldo_base_proporcional"])
    if resumen["valor_horas_extra"]:
        fila_haber(f"Horas Extras ({datos['horas_extra']} hrs × $5.000)",
                   resumen["valor_horas_extra"], 0,
                   resumen["valor_horas_extra"])
    if resumen["bono_colacion"]:
        fila_haber("Bono Colación", 0,
                   resumen["bono_colacion"], resumen["bono_colacion"])
    if resumen["bono_movilizacion"]:
        fila_haber("Bono Movilización", 0,
                   resumen["bono_movilizacion"], resumen["bono_movilizacion"])
    if resumen["bono_asistencia"]:
        fila_haber("Bono Asistencia Perfecta", 0,
                   resumen["bono_asistencia"], resumen["bono_asistencia"])
    if resumen["asig_familiar"]:
        fila_haber("Asignación Familiar", 0,
                   resumen["asig_familiar"], resumen["asig_familiar"])

    # Total haberes
    fila_haber("TOTAL HABERES",
               resumen["total_imponible"],
               resumen["total_haberes"] - resumen["total_imponible"],
               resumen["total_haberes"], bold=True)
    # Forzar color al total
    for col in range(1, 5):
        ws.cell(row=fila-1, column=col).fill = PatternFill("solid", fgColor=C_VERDE_CLA)

    fila += 1  # espacio

    # ─── DESCUENTOS ───────────────────────────
    encabezado_seccion("DESCUENTOS", color_bg=C_ROJO)

    ws.row_dimensions[fila].height = 18
    for col, txt in enumerate(["Concepto", "Tasa", "", "Monto"], 1):
        _celda(ws, fila, col, txt, negrita=True,
               alinear="center", bg="FFFFEBEE", tamanio=9)
    fila += 1

    def fila_desc(concepto, tasa_txt, monto, bold=False):
        nonlocal fila
        ws.row_dimensions[fila].height = 17
        bg = C_BLANCO if fila % 2 == 0 else "FFFFF8F8"
        _celda(ws, fila, 1, concepto, negrita=bold, bg=bg, tamanio=9)
        _celda(ws, fila, 2, tasa_txt, negrita=bold, alinear="center", bg=bg, tamanio=9)
        _celda(ws, fila, 3, "", bg=bg)
        _celda(ws, fila, 4, monto, negrita=bold, alinear="right",
               bg=bg, tamanio=9, formato_num='#,##0')
        fila += 1

    fila_desc(f"AFP ({resumen['tasa_afp']*100:.0f}%)",
              f"{resumen['tasa_afp']*100:.2f}%",
              resumen["desc_afp"])
    fila_desc(f"Salud/Isapre ({resumen['tasa_salud']*100:.0f}%)",
              f"{resumen['tasa_salud']*100:.2f}%",
              resumen["desc_salud"])
    fila_desc(f"Seguro de Cesantia ({resumen['tasa_cesantia']*100:.1f}%)",
              f"{resumen['tasa_cesantia']*100:.2f}%",
              resumen["desc_cesantia"])
    if resumen["desc_prestamo"]:
        fila_desc("Prestamo Empresa",  "—", resumen["desc_prestamo"])
    if resumen["desc_adelanto"]:
        fila_desc("Adelanto de Sueldo","—", resumen["desc_adelanto"])
    if resumen["desc_otros"]:
        fila_desc("Otros Descuentos",  "—", resumen["desc_otros"])

    fila_desc("TOTAL DESCUENTOS", "", resumen["total_descuentos"], bold=True)
    for col in range(1, 5):
        ws.cell(row=fila-1, column=col).fill = PatternFill("solid", fgColor="FFFFCDD2")

    fila += 1  # espacio

    # ─── LÍQUIDO A PAGAR ──────────────────────
    ws.row_dimensions[fila].height = 30
    ws.merge_cells(f"A{fila}:C{fila}")
    c = ws.cell(row=fila, column=1, value="LÍQUIDO A PAGAR")
    c.font = Font(name="Calibri", bold=True, size=14, color=C_BLANCO)
    c.fill = PatternFill("solid", fgColor=C_VERDE_OSC)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = _borde("medium")

    c2 = ws.cell(row=fila, column=4, value=resumen["liquido_a_pagar"])
    c2.font = Font(name="Calibri", bold=True, size=14, color=C_AMARILLO)
    c2.fill = PatternFill("solid", fgColor=C_VERDE_OSC)
    c2.alignment = Alignment(horizontal="right", vertical="center")
    c2.number_format = '#,##0'
    c2.border = _borde("medium")
    fila += 2

    # ─── FIRMA ────────────────────────────────
    ws.row_dimensions[fila].height = 30
    ws.merge_cells(f"A{fila}:A{fila+2}")
    ws.merge_cells(f"C{fila}:D{fila+2}")
    for col, texto in [(1, datos["nombre"]), (3, datos.get("empresa", "Servicarg"))]:
        c = ws.cell(row=fila, column=col, value=texto)
        c.alignment = Alignment(horizontal="center", vertical="bottom")
        c.font = Font(name="Calibri", size=9, color=C_GRIS_TEXTO)
    fila += 3
    ws.merge_cells(f"A{fila}:A{fila}")
    ws.merge_cells(f"C{fila}:D{fila}")
    for col, label in [(1, "Trabajador"), (3, "Empleador / Empresa")]:
        c = ws.cell(row=fila, column=col, value=label)
        c.font = Font(name="Calibri", bold=True, size=9, color=C_VERDE_OSC)
        c.alignment = Alignment(horizontal="center")

    # ─── SEGUNDA HOJA: RESUMEN NUMÉRICO ───────
    ws2 = wb.create_sheet("Resumen Numérico")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 18

    resumen_filas = [
        ("Sueldo base (proporcional)", resumen["sueldo_base_proporcional"]),
        ("Horas extras", resumen["valor_horas_extra"]),
        ("Bono colación", resumen["bono_colacion"]),
        ("Bono movilización", resumen["bono_movilizacion"]),
        ("Bono asistencia", resumen["bono_asistencia"]),
        ("Asignación familiar", resumen["asig_familiar"]),
        ("— Total haberes —", resumen["total_haberes"]),
        ("AFP", -resumen["desc_afp"]),
        ("Salud", -resumen["desc_salud"]),
        ("Cesantía", -resumen["desc_cesantia"]),
        ("Préstamo", -resumen["desc_prestamo"]),
        ("Adelanto", -resumen["desc_adelanto"]),
        ("Otros desc.", -resumen["desc_otros"]),
        ("— Total descuentos —", -resumen["total_descuentos"]),
        ("LÍQUIDO A PAGAR", resumen["liquido_a_pagar"]),
    ]
    ws2.cell(row=1, column=1, value="Concepto").font = Font(bold=True, size=11)
    ws2.cell(row=1, column=2, value="Monto ($)").font = Font(bold=True, size=11)
    for i, (concepto, monto) in enumerate(resumen_filas, start=2):
        ws2.cell(row=i, column=1, value=concepto)
        c = ws2.cell(row=i, column=2, value=monto)
        c.number_format = '#,##0'
        c.alignment = Alignment(horizontal="right")
        if "Total" in concepto or "LÍQUIDO" in concepto:
            ws2.cell(row=i, column=1).font = Font(bold=True)
            c.font = Font(bold=True)

    wb.save(ruta_salida)
